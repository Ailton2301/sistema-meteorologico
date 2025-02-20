import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import requests

def capturar_dados(api_key, cidade):
    url = f"https://api.openweathermap.org/data/2.5/weather?q={cidade}&appid={api_key}&units=metric&lang=pt"
    response = requests.get(url)
    dados = response.json()
    
    if response.status_code != 200:
        raise Exception(dados.get("message", "Erro ao obter dados da API"))
    
    
    temperatura = dados["main"]["temp"]
    sensacao_termica = dados["main"]["feels_like"]
    pressao = dados["main"]["pressure"]
    umidade = dados["main"]["humidity"]
    velocidade_vento = dados["wind"]["speed"]
    data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    return [data_hora, cidade, temperatura, sensacao_termica, pressao, umidade, velocidade_vento]

def salvar_dados(file_path, dados):
    try:
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Data/Hora", "Cidade", "Temperatura (°C)", "Sensação Térmica (°C)", "Pressão (hPa)", "Umidade (%)", "Velocidade do Vento (m/s)"])
        
        
        sheet.append(dados)
        workbook.save(file_path)
        print(f"Arquivo salvo em: {file_path}")
    except PermissionError:
        raise Exception("Erro: Permissão negada. Verifique se o arquivo está aberto ou se você tem permissão para escrever na pasta.")
    except Exception as e:
        raise Exception(f"Erro ao salvar o arquivo: {e}")

def executar_captura():
    API_KEY = "7345c6c368262bfe468714ef9872bd33" # Substituir pela sua chave de API
    CITY = entry_cidade.get().strip()
    FILE_NAME = "dados_tempo.xlsx"
    FILE_PATH = os.path.join("C:\\", "Users", "silva", "OneDrive", "Documentos", "estudos", "python", "portifolio", FILE_NAME)  # Caminho completo
    
    if not CITY:
        messagebox.showwarning("Aviso", "Por favor, insira o nome da cidade.")
        return
    
    try:
        dados = capturar_dados(API_KEY, CITY)
        salvar_dados(FILE_PATH, dados)
        messagebox.showinfo("Sucesso", f"Dados capturados e salvos com sucesso!\nArquivo: {FILE_PATH}")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

def criar_interface():
    global entry_cidade
    
    root = tk.Tk()
    root.title("Captura de Dados Meteorológicos")
    
    label_cidade = tk.Label(root, text="Cidade:")
    label_cidade.pack(pady=5)
    entry_cidade = tk.Entry(root, width=30)
    entry_cidade.pack(pady=5)
    
    btn_capturar = tk.Button(root, text="Capturar Dados", command=executar_captura)
    btn_capturar.pack(pady=20)
    
    root.mainloop()

if __name__ == "__main__":
    criar_interface()  